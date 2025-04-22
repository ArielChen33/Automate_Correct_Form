import pandas as pd
import math
from datetime import datetime
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Sellect input files
Tk().withdraw()  # Hide root window
file_path = filedialog.askopenfilename(
    title="Select your Grocery Purchase Order file",
    filetypes=[("Excel files", "*.xlsx")]
)

if not file_path: 
    print("No file selected. Exiting.")
    exit()


print("You selected:", file_path)

# Create base name for output
base_name = file_path.split("/")[-1].replace("Purchase Order", "Shipping Deteil").replace("xlsx", "")
date_str = datetime.now().strftime("%m%d-%Y")
output_file = f"{base_name} - {date_str}.xlsx"

# Load the excel file
df = pd.read_excel(file_path)
output_rows = []

# Iterate over each columns name in the excel file
for index, row in df.iterrows():
    product_name = row["product name"] 
    total_quantity = int(row["total quantity"])
    unit = int(row["unit"])

    # calculate the numbers of rows
    row_quantity = math.ceil(total_quantity / unit)

    for _ in range(row_quantity):
        output_rows.append({
            "product name": product_name, 
            "total quantity": total_quantity, 
            "unit": unit, 
            "UCC label": "", 
            "tracking number": ""
        })

    # Add blank row after each product block
    output_rows.append({
        "product name": product_name, 
        "total quantity": total_quantity, 
        "unit": unit, 
        "UCC label": "", 
        "tracking number": ""
    })

# Export yo new excel file
output_df = pd.DataFrame(output_rows)
output_df.to_excel(output_file, index=False)

# Apply color grouping
colors = ["FFEBEE", "E8F5E9", "E3F2FD", "FFF3E0", "F3E5F5", "E0F7FA", "F9FBE7"]

wb = load_workbook(output_file)
ws = wb.active # select the first sheet in the workbook

header_row = [cell.value for cell in ws[1]]
product_col_index = header_row.index("product name") + 1  # openpyxl is 1-indexed

colors = ["FFEBEE", "E8F5E9", "E3F2FD", "FFF3E0", "F3E5F5", "E0F7FA", "F9FBE7"]
color_index = 0
prev_product = None

for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
    curr_product = ws.cell(row=row, column=product_col_index).value

    if curr_product != prev_product:
        color_index = (color_index + 1) % len(colors)
        prev_product = curr_product

    fill_color = colors[color_index]

    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color=fill_color,
            end_color=fill_color,
            fill_type="solid"
        )


wb.save(output_file)


print(f"Export completed: {output_file}")
