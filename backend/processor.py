import pandas as pd
import math
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def process_excel(input_path):
    df = pd.read_excel(input_path)
    base_name = input_path.split("/")[-1].replace("Purchase Order", "Shipping Detail").replace(".xlsx", "")
    date_str = datetime.now().strftime("%m%d-%Y")
    output_path = f"outputs/{base_name} - {date_str}.xlsx"

    colors = [
        "FFEBEE", "E8F5E9", "E3F2FD",
        "FFF3E0", "F3E5F5", "E0F7FA", "F9FBE7"
    ]

    # Expand rows
    output_rows = []
    for _, row in df.iterrows():
        total = int(row["total quantity"])
        unit = int(row["unit"])
        product = row["product name"]
        rows = math.ceil(total / unit)
        for _ in range(rows):
            output_rows.append({
                "product name": product,
                "total quantity": total,
                "unit": unit,
                "UCC label": "",
                "tracking number": ""
            })
        # Add blank row
        output_rows.append({
            "product name": product,
            "total quantity": total,
            "unit": unit,
            "UCC label": "",
            "tracking number": ""
        })

    # Export to Excel
    out_df = pd.DataFrame(output_rows)
    out_df.to_excel(output_path, index=False)

    # Apply color grouping by product name
    wb = load_workbook(output_path)
    ws = wb.active

    # Find column index of "product name"
    header_row = [cell.value for cell in ws[1]]
    product_col_index = header_row.index("product name") + 1  # 1-based index

    color_index = 0
    prev_product = None

    for row in range(2, ws.max_row + 1):  # start after header
        curr_product = ws.cell(row=row, column=product_col_index).value

        if curr_product != prev_product:
            color_index = (color_index + 1) % len(colors)
            prev_product = curr_product

        fill_color = colors[color_index]

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid"
            )

    wb.save(output_path)
    return output_path


# import pandas as pd
# import math
# from datetime import datetime
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill

# def process_excel(input_path):
#     df = pd.read_excel(input_path)
#     base_name = input_path.split("/")[-1].replace("Purchase Order", "Shipping Detail").replace(".xlsx", "")
#     date_str = datetime.now().strftime("%m%d-%Y")
#     output_path = f"outputs/{base_name} - {date_str}.xlsx"

#     colors = [
#         "FFEBEE", "E8F5E9", "E3F2FD",
#         "FFF3E0", "F3E5F5", "E0F7FA", "F9FBE7"
#     ]

#     # Expand rows
#     output_rows = []
#     for _, row in df.iterrows():
#         total = int(row["total quantity"])
#         unit = int(row["unit"])
#         rows = math.ceil(total / unit)
#         for _ in range(rows):
#             output_rows.append({
#                 "product name": row["product name"],
#                 "total quantity": total,
#                 "unit": unit,
#                 "UCC label": "",
#                 "tracking number": ""
#             })
#         output_rows.append({k: "" for k in output_rows[-1].keys()})  # blank row

#     out_df = pd.DataFrame(output_rows)
#     out_df.to_excel(output_path, index=False)

#     wb = load_workbook(output_path)
#     ws = wb.active
#     for row in range(2, ws.max_row + 1):
#         color_index = (row - 2) % len(colors)
#         fill_color = colors[color_index]
#         for col in range(1, ws.max_column + 1):
#             ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
#     wb.save(output_path)
#     return output_path
